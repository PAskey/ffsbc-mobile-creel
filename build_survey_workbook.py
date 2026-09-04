#!/usr/bin/env python3
"""
build_survey_workbook.py
------------------------
Turn a batch of Marquart Lake mobile-app catch reports (JSON exported from the
angler PWA) into a filled copy of the FFSBC SURVEY.xlsx upload template.

What it fills:
  - Party  : one row per submission (trip-level record)
  - Count  : one row per species caught (num_caught / num_kept)
Left blank on purpose: Angler, IndividualFish, ExtendedSurvey.
Untouched: SubmissionDetail, AssessEvent, and all reference tabs
           (Codes, Species, Waterbodies, Projects).

Constants that live on Party row 2 of the template (survey_type, trip_complete,
location_description, ...) are read and used as defaults, so anything the app
omits still falls back to the value you pre-filled.

Usage:
  python build_survey_workbook.py SUBMISSIONS.json --template SURVEY.xlsx --out SURVEY_filled.xlsx

Prototype note: openpyxl re-saves the whole workbook and drops Excel's
data-validation / conditional-formatting *extensions* on the green tabs. For the
real upload we should validate the output against the FFSBC submission site, or
do this same fill in R (openxlsx preserves the template more faithfully).
"""
import argparse, json, sys
from collections import defaultdict
import openpyxl


def header_map(ws, header_row=1):
    """Return {normalized_header: column_index} for a sheet's header row."""
    m = {}
    for col, cell in enumerate(ws[header_row], start=1):
        if cell.value is not None:
            m[str(cell.value).strip()] = col
    return m


def read_row_as_dict(ws, hdr, row_idx):
    return {name: ws.cell(row=row_idx, column=col).value for name, col in hdr.items()}


def clear_data_rows(ws, header_row=1, first_data_row=2):
    """Blank out any existing values below the header so we start clean."""
    if ws.max_row < first_data_row:
        return
    for row in ws.iter_rows(min_row=first_data_row, max_row=ws.max_row):
        for cell in row:
            cell.value = None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("submissions", help="JSON file exported from the app")
    ap.add_argument("--template", default="SURVEY.xlsx")
    ap.add_argument("--out", default="SURVEY_filled.xlsx")
    ap.add_argument("--count-zero-for-no-catch", action="store_true",
                    help="Also write a Count row (0 caught / 0 kept) for skunked trips.")
    args = ap.parse_args()

    with open(args.submissions, encoding="utf-8") as f:
        subs = json.load(f)
    if isinstance(subs, dict):
        subs = subs.get("reports") or subs.get("submissions") or [subs]

    # FFSBC party_id / person_id must be integers 1..500, and the tally restarts each
    # day at a lake. Assign party_id here (the app can't, since phones can't coordinate
    # numbers): sort by date+time and number 1,2,3... within each interview_date.
    subs = sorted(subs, key=lambda s: ((s.get("interview_date") or ""),
                                       (s.get("interview_time") or ""),
                                       (s.get("client_ref") or "")))
    day_seq = defaultdict(int)
    over_limit = set()
    for s in subs:
        d = s.get("interview_date") or ""
        day_seq[d] += 1
        s["_pid"] = day_seq[d]
        if day_seq[d] > 500:
            over_limit.add(d)

    wb = openpyxl.load_workbook(args.template)
    party = wb["Party"]
    count = wb["Count"]
    angler = wb["Angler"]

    p_hdr = header_map(party)
    c_hdr = header_map(count)
    a_hdr = header_map(angler)

    # Defaults from the constants row you pre-filled (Party row 2).
    p_defaults = read_row_as_dict(party, p_hdr, 2)

    # Fresh start on the data tabs (leaves headers + all other tabs intact).
    clear_data_rows(party)
    clear_data_rows(count)
    clear_data_rows(angler)

    def pset(row_idx, name, value):
        if name in p_hdr and value is not None and value != "":
            party.cell(row=row_idx, column=p_hdr[name]).value = value

    def cset(row_idx, name, value):
        if name in c_hdr and value is not None and value != "":
            count.cell(row=row_idx, column=c_hdr[name]).value = value

    def aset(row_idx, name, value):
        if name in a_hdr and value is not None and value != "":
            angler.cell(row=row_idx, column=a_hdr[name]).value = value

    p_row = 2
    c_row = 2
    a_row = 2
    n_party = 0
    n_count = 0
    n_angler = 0

    for s in subs:
        # ---- Party row -------------------------------------------------
        def val(key):
            v = s.get(key)
            return v if v not in (None, "") else p_defaults.get(key)

        pset(p_row, "assessment_event_name", val("assessment_event_name"))
        pset(p_row, "survey_type",           val("survey_type"))
        pset(p_row, "interview_date",         s.get("interview_date"))
        pset(p_row, "party_id",               s["_pid"])
        pset(p_row, "interview_time",         s.get("interview_time"))
        pset(p_row, "trip_complete",          val("trip_complete"))
        pset(p_row, "location_description",   val("location_description"))
        pset(p_row, "fishing_mode",           s.get("fishing_mode"))
        pset(p_row, "num_anglers",            s.get("num_anglers"))
        pset(p_row, "num_rods",               s.get("num_rods"))
        pset(p_row, "hours_fished",           s.get("hours_fished"))
        pset(p_row, "target_species_1",       s.get("target_species_1"))
        pset(p_row, "target_species_2",       s.get("target_species_2"))
        # angling_method / terminal_gear are collected per respondent -> Angler tab,
        # so Party's copies of those columns are left blank on purpose.
        pset(p_row, "comment",                s.get("comment"))
        p_row += 1
        n_party += 1

        person_id = s.get("person_id") or 1
        aen = val("assessment_event_name")
        pid = s["_pid"]

        # ---- Angler row (the reporting angler) -------------------------
        if any(s.get(k) not in (None, "") for k in ("angling_method", "terminal_gear", "years_fishing")):
            aset(a_row, "assessment_event_name", aen)
            aset(a_row, "interview_date",        s.get("interview_date"))
            aset(a_row, "party_id",              pid)
            aset(a_row, "person_id",             person_id)
            aset(a_row, "angling_method",        s.get("angling_method"))
            aset(a_row, "terminal_gear",         s.get("terminal_gear"))
            aset(a_row, "years_fishing",         s.get("years_fishing"))
            a_row += 1
            n_angler += 1

        # ---- Count rows ------------------------------------------------
        catches = s.get("catches") or []
        if not catches and s.get("no_catch") and args.count_zero_for_no_catch:
            catches = [{"species": None, "num_caught": 0, "num_kept": 0}]
        for c in catches:
            cset(c_row, "assessment_event_name", aen)
            cset(c_row, "interview_date",        s.get("interview_date"))
            cset(c_row, "party_id",              pid)
            cset(c_row, "person_id",             person_id)
            cset(c_row, "species",               c.get("species"))
            cset(c_row, "num_caught",            c.get("num_caught"))
            cset(c_row, "num_kept",              c.get("num_kept"))
            c_row += 1
            n_count += 1

    wb.save(args.out)
    print(f"Wrote {args.out}: {n_party} Party, {n_angler} Angler, {n_count} Count row(s) "
          f"from {len(subs)} submission(s).")
    print("party_id assigned 1..N per interview_date: " +
          ", ".join(f"{d}={n}" for d, n in sorted(day_seq.items())))
    if over_limit:
        print(f"WARNING: more than 500 reports on {sorted(over_limit)} — party_id exceeds "
              f"the 1–500 limit for those days; split the upload by day.")


if __name__ == "__main__":
    main()
